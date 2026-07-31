"""Conciliación contra el libro auxiliar de contabilidad.

El resumen por banco dice qué pasó en el banco. La conciliación propiamente
dicha compara eso con lo que registró la contabilidad y explica la diferencia:

    Saldo del extracto  + partidas registradas en libros que el banco aún no
                          refleja (cheques girados sin cobrar, consignaciones
                          en tránsito)
    Saldo en libros     + partidas del banco no registradas en libros
                          (GMF, comisiones, notas débito/crédito)
    -----------------------------------------------------------------
    Ambos caminos deben llegar al mismo saldo conciliado.

Convención de signos del libro: positivo = entra plata al banco (débito en la
cuenta del banco, que es un activo); negativo = sale plata.
"""

from __future__ import annotations

import csv
import os
import re
from dataclasses import dataclass, field
from datetime import date
from decimal import Decimal

from .modelos import Movimiento
from .normalizacion import CENTAVOS, clave, formato_cop, parse_fecha, parse_monto
from .reportes import DERECHA, IZQUIERDA, tabla

CERO = Decimal("0.00")

ENCABEZADOS = {
    "fecha": ("FECHA", "FEC", "DIA", "DATE", "FECHA MOVIMIENTO", "FECHA DOCUMENTO"),
    "descripcion": (
        "DESCRIPCION", "DETALLE", "CONCEPTO", "GLOSA", "NOTA", "OBSERVACION",
        "TERCERO", "BENEFICIARIO", "NOMBRE",
    ),
    "valor": ("VALOR", "MONTO", "IMPORTE", "NETO", "VALOR NETO", "SALDO MOVIMIENTO"),
    "debito": ("DEBITO", "DEBITOS", "DEBE", "CARGO", "CARGOS", "VALOR DEBITO"),
    "credito": ("CREDITO", "CREDITOS", "HABER", "ABONO", "ABONOS", "VALOR CREDITO"),
    "referencia": (
        "REFERENCIA", "DOCUMENTO", "COMPROBANTE", "CHEQUE", "NUMERO", "NRO",
        "CONSECUTIVO", "SOPORTE",
    ),
    "banco": ("BANCO", "ENTIDAD", "CUENTA BANCARIA"),
    "cuenta": ("CUENTA", "NUMERO CUENTA", "CTA"),
}


@dataclass
class ApunteLibro:
    """Un registro del libro auxiliar."""

    fecha: date
    descripcion: str
    valor: Decimal
    referencia: str | None = None
    banco: str | None = None
    cuenta: str | None = None
    fila: int = 0

    @property
    def tipo(self) -> str:
        return "INGRESO" if self.valor > 0 else "EGRESO"


# ---------------------------------------------------------------------------
# Lectura del archivo
# ---------------------------------------------------------------------------

def _mapear_columnas(encabezado: list[str]) -> dict[str, int]:
    """Relaciona los títulos del archivo con los campos que necesitamos."""
    mapa: dict[str, int] = {}
    normalizados = [clave(c) for c in encabezado]
    for campo, alias in ENCABEZADOS.items():
        for indice, titulo in enumerate(normalizados):
            if not titulo or indice in mapa.values():
                continue
            if titulo in alias or any(
                titulo.startswith(a) or a in titulo for a in alias
            ):
                mapa[campo] = indice
                break
    return mapa


def _filas_de_csv(ruta: str) -> list[list[str]]:
    with open(ruta, "r", encoding="utf-8-sig", errors="replace") as archivo:
        muestra = archivo.read(4096)
        archivo.seek(0)
        try:
            dialecto = csv.Sniffer().sniff(muestra, delimiters=";,\t|")
        except csv.Error:
            dialecto = csv.excel
            dialecto.delimiter = ";" if muestra.count(";") > muestra.count(",") else ","
        return [fila for fila in csv.reader(archivo, dialecto) if any(fila)]


def _filas_de_excel(ruta: str) -> list[list[str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as error:  # pragma: no cover
        raise RuntimeError(
            "Para leer archivos de Excel se necesita openpyxl. "
            "Instala con: pip install openpyxl"
        ) from error

    libro = load_workbook(ruta, data_only=True, read_only=True)
    hoja = libro.active
    filas: list[list[str]] = []
    for fila in hoja.iter_rows(values_only=True):
        if fila is None or all(celda is None for celda in fila):
            continue
        filas.append(["" if celda is None else str(celda) for celda in fila])
    libro.close()
    return filas


def leer_libro(
    ruta: str, *, invertir_signo: bool = False, anio_defecto: int | None = None
) -> tuple[list[ApunteLibro], list[str]]:
    """Lee el libro auxiliar desde CSV o Excel.

    Detecta solo las columnas que reconoce por su título; si el archivo no trae
    encabezados, asume el orden fecha, descripción, valor.
    """
    extension = os.path.splitext(ruta)[1].lower()
    filas = _filas_de_excel(ruta) if extension in (".xlsx", ".xlsm") else _filas_de_csv(ruta)
    if not filas:
        return [], ["El libro auxiliar está vacío."]

    advertencias: list[str] = []
    mapa = _mapear_columnas(filas[0])
    tiene_encabezado = "fecha" in mapa and (
        "valor" in mapa or "debito" in mapa or "credito" in mapa
    )
    if not tiene_encabezado:
        mapa = {"fecha": 0, "descripcion": 1, "valor": 2}
        advertencias.append(
            "No reconocí los títulos de las columnas: asumo orden "
            "fecha, descripción, valor."
        )
        cuerpo = filas
    else:
        cuerpo = filas[1:]

    apuntes: list[ApunteLibro] = []
    descartadas = 0

    for numero, fila in enumerate(cuerpo, start=2 if tiene_encabezado else 1):
        def celda(campo: str) -> str:
            indice = mapa.get(campo)
            if indice is None or indice >= len(fila):
                return ""
            return (fila[indice] or "").strip()

        fecha = parse_fecha(celda("fecha"), anio_defecto)
        if fecha is None:
            descartadas += 1
            continue

        if "valor" in mapa and celda("valor"):
            valor = parse_monto(celda("valor"))
        else:
            debito = parse_monto(celda("debito")) or CERO
            credito = parse_monto(celda("credito")) or CERO
            valor = abs(debito) - abs(credito)

        if valor is None or valor == 0:
            descartadas += 1
            continue

        if invertir_signo:
            valor = -valor

        apuntes.append(
            ApunteLibro(
                fecha=fecha,
                descripcion=celda("descripcion") or "(sin descripción)",
                valor=valor.quantize(CENTAVOS),
                referencia=celda("referencia") or None,
                banco=celda("banco") or None,
                cuenta=celda("cuenta") or None,
                fila=numero,
            )
        )

    if descartadas:
        advertencias.append(
            f"Se omitieron {descartadas} filas del libro sin fecha o sin valor."
        )
    return apuntes, advertencias


# ---------------------------------------------------------------------------
# Emparejamiento
# ---------------------------------------------------------------------------

@dataclass
class Pareja:
    movimiento: Movimiento
    apunte: ApunteLibro

    @property
    def dias(self) -> int:
        return abs((self.movimiento.fecha - self.apunte.fecha).days)


@dataclass
class ResultadoConciliacion:
    parejas: list[Pareja] = field(default_factory=list)
    solo_banco: list[Movimiento] = field(default_factory=list)
    solo_libro: list[ApunteLibro] = field(default_factory=list)
    saldo_extracto: Decimal | None = None
    saldo_libros: Decimal | None = None
    advertencias: list[str] = field(default_factory=list)

    @property
    def ajuste_desde_banco(self) -> Decimal:
        """Partidas del libro que el banco todavía no refleja."""
        return sum((a.valor for a in self.solo_libro), CERO).quantize(CENTAVOS)

    @property
    def ajuste_desde_libros(self) -> Decimal:
        """Partidas del banco que la contabilidad no registró."""
        return sum((m.valor for m in self.solo_banco), CERO).quantize(CENTAVOS)

    @property
    def saldo_conciliado_banco(self) -> Decimal | None:
        if self.saldo_extracto is None:
            return None
        return (self.saldo_extracto + self.ajuste_desde_banco).quantize(CENTAVOS)

    @property
    def saldo_conciliado_libros(self) -> Decimal | None:
        if self.saldo_libros is None:
            return None
        return (self.saldo_libros + self.ajuste_desde_libros).quantize(CENTAVOS)

    @property
    def diferencia(self) -> Decimal | None:
        a, b = self.saldo_conciliado_banco, self.saldo_conciliado_libros
        if a is None or b is None:
            return None
        return (a - b).quantize(CENTAVOS)

    @property
    def concilia(self) -> bool | None:
        diferencia = self.diferencia
        return None if diferencia is None else abs(diferencia) <= Decimal("0.02")


def _mismo_banco(movimiento: Movimiento, apunte: ApunteLibro) -> bool:
    """Si el libro trae banco, se exige que coincida; si no, no se filtra."""
    if not apunte.banco:
        return True
    izquierda, derecha = clave(movimiento.banco), clave(apunte.banco)
    if izquierda in derecha or derecha in izquierda:
        return True
    # "BANCOLOMBIA AHORROS" contra "Bancolombia": basta la primera palabra.
    return derecha.split()[0] in izquierda if derecha.split() else False


def conciliar(
    movimientos: list[Movimiento],
    apuntes: list[ApunteLibro],
    *,
    dias_tolerancia: int = 5,
    saldo_extracto: Decimal | None = None,
    saldo_libros: Decimal | None = None,
) -> ResultadoConciliacion:
    """Empareja movimientos del banco con apuntes del libro.

    Un par válido tiene el mismo valor (signo incluido) y fechas dentro de la
    tolerancia. Se prefiere siempre la fecha más cercana, y cada apunte se usa
    una sola vez, para que dos pagos iguales no se emparejen con el mismo
    registro.
    """
    resultado = ResultadoConciliacion(
        saldo_extracto=saldo_extracto, saldo_libros=saldo_libros
    )

    disponibles: dict[Decimal, list[ApunteLibro]] = {}
    for apunte in apuntes:
        disponibles.setdefault(apunte.valor, []).append(apunte)

    usados: set[int] = set()

    for movimiento in sorted(movimientos, key=lambda m: m.fecha):
        candidatos = [
            a
            for a in disponibles.get(movimiento.valor, [])
            if id(a) not in usados
            and abs((movimiento.fecha - a.fecha).days) <= dias_tolerancia
            and _mismo_banco(movimiento, a)
        ]
        if not candidatos:
            resultado.solo_banco.append(movimiento)
            continue
        # Primero la fecha más cercana; si empatan, la referencia parecida.
        elegido = min(
            candidatos,
            key=lambda a: (
                abs((movimiento.fecha - a.fecha).days),
                0 if _referencias_parecidas(movimiento, a) else 1,
                a.fila,
            ),
        )
        usados.add(id(elegido))
        resultado.parejas.append(Pareja(movimiento=movimiento, apunte=elegido))

    resultado.solo_libro = [a for a in apuntes if id(a) not in usados]

    _revisar_signos(resultado, movimientos, apuntes)
    return resultado


def _referencias_parecidas(movimiento: Movimiento, apunte: ApunteLibro) -> bool:
    referencia = re.sub(r"\D", "", apunte.referencia or "")
    if not referencia or len(referencia) < 3:
        return False
    candidatos = " ".join(
        filter(None, [movimiento.documento, movimiento.referencia, movimiento.descripcion])
    )
    return referencia.lstrip("0") in re.sub(r"\D", " ", candidatos)


def _revisar_signos(
    resultado: ResultadoConciliacion,
    movimientos: list[Movimiento],
    apuntes: list[ApunteLibro],
) -> None:
    """Avisa si el libro parece tener la convención de signos al revés."""
    if not movimientos or not apuntes:
        return
    if len(resultado.parejas) >= len(movimientos) * 0.5:
        return

    valores_libro = {-a.valor for a in apuntes}
    coincidencias_invertidas = sum(1 for m in movimientos if m.valor in valores_libro)
    if coincidencias_invertidas > len(resultado.parejas):
        resultado.advertencias.append(
            "Casi no hubo coincidencias, pero sí las hay al invertir el signo: "
            "tu libro probablemente usa débito/crédito al contrario. "
            "Vuelve a correrlo con --invertir-signo."
        )


# ---------------------------------------------------------------------------
# Reporte
# ---------------------------------------------------------------------------

def reporte_conciliacion(resultado: ResultadoConciliacion) -> str:
    bloques: list[str] = []

    bloques.append("PARTIDAS CONCILIADAS")
    bloques.append(
        tabla(
            ["Fecha banco", "Descripción banco", "Valor", "Fecha libro",
             "Descripción libro", "Días"],
            [
                [
                    p.movimiento.fecha.strftime("%d/%m/%Y"),
                    p.movimiento.descripcion[:34],
                    formato_cop(p.movimiento.valor),
                    p.apunte.fecha.strftime("%d/%m/%Y"),
                    p.apunte.descripcion[:34],
                    str(p.dias),
                ]
                for p in sorted(resultado.parejas, key=lambda p: p.movimiento.fecha)
            ],
            [IZQUIERDA, IZQUIERDA, DERECHA, IZQUIERDA, IZQUIERDA, DERECHA],
        )
    )

    bloques.append("")
    bloques.append(
        "PARTIDAS DEL BANCO NO REGISTRADAS EN LIBROS "
        "(comisiones, GMF, notas débito/crédito)"
    )
    bloques.append(
        tabla(
            ["Fecha", "Banco", "Descripción", "Valor"],
            [
                [
                    m.fecha.strftime("%d/%m/%Y"),
                    m.banco,
                    m.descripcion[:48],
                    formato_cop(m.valor),
                ]
                for m in sorted(resultado.solo_banco, key=lambda m: m.fecha)
            ],
            [IZQUIERDA, IZQUIERDA, IZQUIERDA, DERECHA],
        )
    )

    bloques.append("")
    bloques.append(
        "PARTIDAS DE LIBROS NO REFLEJADAS EN EL BANCO "
        "(cheques sin cobrar, consignaciones en tránsito)"
    )
    bloques.append(
        tabla(
            ["Fecha", "Descripción", "Referencia", "Valor"],
            [
                [
                    a.fecha.strftime("%d/%m/%Y"),
                    a.descripcion[:48],
                    a.referencia or "-",
                    formato_cop(a.valor),
                ]
                for a in sorted(resultado.solo_libro, key=lambda a: a.fecha)
            ],
            [IZQUIERDA, IZQUIERDA, IZQUIERDA, DERECHA],
        )
    )

    bloques.append("")
    bloques.append("CUADRE DE LA CONCILIACIÓN")
    filas = [
        ["Saldo según extracto bancario", formato_cop(resultado.saldo_extracto)],
        ["(+/-) Partidas de libros pendientes en el banco",
         formato_cop(resultado.ajuste_desde_banco)],
        ["= Saldo conciliado (vía banco)",
         formato_cop(resultado.saldo_conciliado_banco)],
        ["", ""],
        ["Saldo según libros", formato_cop(resultado.saldo_libros)],
        ["(+/-) Partidas del banco no registradas",
         formato_cop(resultado.ajuste_desde_libros)],
        ["= Saldo conciliado (vía libros)",
         formato_cop(resultado.saldo_conciliado_libros)],
    ]
    diferencia = resultado.diferencia
    if diferencia is not None:
        filas.append(["", ""])
        filas.append(["DIFERENCIA", formato_cop(diferencia)])
        filas.append(
            ["Estado", "CONCILIADO" if resultado.concilia else "NO CONCILIA"]
        )
    bloques.append(tabla(["Concepto", "Valor"], filas, [IZQUIERDA, DERECHA]))

    resumen = [
        f"Conciliadas: {len(resultado.parejas)}",
        f"Solo banco: {len(resultado.solo_banco)}",
        f"Solo libros: {len(resultado.solo_libro)}",
    ]
    bloques.append("")
    bloques.append("  ·  ".join(resumen))

    if resultado.saldo_libros is None:
        bloques.append(
            "\nNota: no se indicó el saldo en libros (--saldo-libros), así que "
            "solo se calculó el lado del banco."
        )

    return "\n".join(bloques)


CAMPOS_CONCILIACION = [
    "estado", "fecha_banco", "banco", "descripcion_banco", "valor_banco",
    "fecha_libro", "descripcion_libro", "referencia_libro", "valor_libro",
    "dias_diferencia",
]


def filas_conciliacion(resultado: ResultadoConciliacion) -> list[dict[str, object]]:
    filas: list[dict[str, object]] = []
    for pareja in resultado.parejas:
        filas.append(
            {
                "estado": "CONCILIADO",
                "fecha_banco": pareja.movimiento.fecha.isoformat(),
                "banco": pareja.movimiento.banco,
                "descripcion_banco": pareja.movimiento.descripcion,
                "valor_banco": pareja.movimiento.valor,
                "fecha_libro": pareja.apunte.fecha.isoformat(),
                "descripcion_libro": pareja.apunte.descripcion,
                "referencia_libro": pareja.apunte.referencia or "",
                "valor_libro": pareja.apunte.valor,
                "dias_diferencia": pareja.dias,
            }
        )
    for movimiento in resultado.solo_banco:
        filas.append(
            {
                "estado": "SOLO BANCO",
                "fecha_banco": movimiento.fecha.isoformat(),
                "banco": movimiento.banco,
                "descripcion_banco": movimiento.descripcion,
                "valor_banco": movimiento.valor,
                "fecha_libro": "",
                "descripcion_libro": "",
                "referencia_libro": "",
                "valor_libro": "",
                "dias_diferencia": "",
            }
        )
    for apunte in resultado.solo_libro:
        filas.append(
            {
                "estado": "SOLO LIBROS",
                "fecha_banco": "",
                "banco": apunte.banco or "",
                "descripcion_banco": "",
                "valor_banco": "",
                "fecha_libro": apunte.fecha.isoformat(),
                "descripcion_libro": apunte.descripcion,
                "referencia_libro": apunte.referencia or "",
                "valor_libro": apunte.valor,
                "dias_diferencia": "",
            }
        )
    return filas
