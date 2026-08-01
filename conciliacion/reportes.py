"""Presentación de resultados: consola, CSV y Excel."""

from __future__ import annotations

import csv
import os
import re
from datetime import datetime
from decimal import Decimal

from .modelos import Extracto, Movimiento, ResumenMensual
from .normalizacion import formato_cop, periodo_legible
from .resumen import Consolidado, TotalBanco, totales_por_mes

DERECHA = "d"
IZQUIERDA = "i"


# ---------------------------------------------------------------------------
# Tablas de consola
# ---------------------------------------------------------------------------

def tabla(
    encabezados: list[str],
    filas: list[list[str]],
    alineaciones: list[str] | None = None,
) -> str:
    """Dibuja una tabla de texto con columnas alineadas."""
    if not filas:
        return "  (sin datos)"
    alineaciones = alineaciones or [IZQUIERDA] * len(encabezados)
    anchos = [len(h) for h in encabezados]
    for fila in filas:
        for i, celda in enumerate(fila):
            anchos[i] = max(anchos[i], len(str(celda)))

    def formatear(valores: list[str]) -> str:
        partes = []
        for i, celda in enumerate(valores):
            texto = str(celda)
            partes.append(
                texto.rjust(anchos[i])
                if alineaciones[i] == DERECHA
                else texto.ljust(anchos[i])
            )
        return "  ".join(partes).rstrip()

    separador = "  ".join("-" * ancho for ancho in anchos)
    lineas = [formatear(encabezados), separador]
    lineas.extend(formatear(fila) for fila in filas)
    return "\n".join(lineas)


def _marca_cuadre(resumen: ResumenMensual) -> str:
    if resumen.saldo_inicial_deducido:
        return "s/verificar"
    cuadra = resumen.cuadra
    if cuadra is None:
        return "s/saldos"
    if cuadra:
        return "OK"
    return f"DIF {formato_cop(resumen.diferencia)}"


def reporte_resumen_mensual(resumenes: list[ResumenMensual]) -> str:
    """El reporte principal: un renglón por banco, cuenta y mes."""
    filas = []
    for resumen in sorted(resumenes, key=lambda r: (r.periodo, r.banco)):
        filas.append(
            [
                periodo_legible(resumen.periodo),
                resumen.banco,
                resumen.cuenta or "-",
                str(resumen.cantidad_movimientos),
                formato_cop(resumen.saldo_inicial),
                formato_cop(resumen.ingresos),
                formato_cop(resumen.egresos),
                formato_cop(resumen.neto),
                formato_cop(resumen.saldo_final),
                _marca_cuadre(resumen),
            ]
        )
    encabezados = [
        "Mes", "Banco", "Cuenta", "Mov.", "Saldo inicial", "Ingresos",
        "Egresos", "Neto", "Saldo final", "Cuadre",
    ]
    alineaciones = [IZQUIERDA] * 3 + [DERECHA] * 6 + [IZQUIERDA]
    return tabla(encabezados, filas, alineaciones)


def reporte_totales_banco(totales: list[TotalBanco]) -> str:
    filas = []
    for total in totales:
        filas.append(
            [
                total.banco,
                ", ".join(total.cuentas) or "-",
                f"{len(total.meses)}",
                str(total.cantidad_movimientos),
                formato_cop(total.saldo_inicial),
                formato_cop(total.ingresos),
                formato_cop(total.egresos),
                formato_cop(total.neto),
                formato_cop(total.saldo_final),
            ]
        )
    if filas:
        filas.append(
            [
                "TOTAL",
                "",
                "",
                str(sum(t.cantidad_movimientos for t in totales)),
                "",
                formato_cop(sum((t.ingresos for t in totales), Decimal("0.00"))),
                formato_cop(sum((t.egresos for t in totales), Decimal("0.00"))),
                formato_cop(sum((t.neto for t in totales), Decimal("0.00"))),
                formato_cop(
                    sum(
                        (t.saldo_final for t in totales if t.saldo_final is not None),
                        Decimal("0.00"),
                    )
                ),
            ]
        )
    encabezados = [
        "Banco", "Cuenta(s)", "Meses", "Mov.", "Saldo inicial", "Ingresos",
        "Egresos", "Neto", "Saldo final",
    ]
    return tabla(encabezados, filas, [IZQUIERDA] * 3 + [DERECHA] * 6)


def reporte_mensual_consolidado(resumenes: list[ResumenMensual]) -> str:
    filas = []
    for fila in totales_por_mes(resumenes):
        filas.append(
            [
                periodo_legible(str(fila["periodo"])),
                str(fila["bancos"]),
                str(fila["cantidad_movimientos"]),
                formato_cop(fila["ingresos"]),
                formato_cop(fila["egresos"]),
                formato_cop(fila["neto"]),
                formato_cop(fila["saldo_final"]),
            ]
        )
    encabezados = [
        "Mes", "Bancos", "Mov.", "Ingresos", "Egresos", "Neto",
        "Saldo final sumado",
    ]
    return tabla(encabezados, filas, [IZQUIERDA] + [DERECHA] * 6)


def reporte_movimientos(movimientos: list[Movimiento], limite: int | None = None) -> str:
    seleccion = movimientos if limite is None else movimientos[:limite]
    filas = []
    for movimiento in seleccion:
        filas.append(
            [
                movimiento.fecha.strftime("%d/%m/%Y"),
                movimiento.banco,
                movimiento.descripcion[:45],
                formato_cop(movimiento.ingreso) if movimiento.ingreso else "",
                formato_cop(movimiento.egreso) if movimiento.egreso else "",
                formato_cop(movimiento.saldo),
                movimiento.confianza.value,
            ]
        )
    texto = tabla(
        ["Fecha", "Banco", "Descripción", "Ingreso", "Egreso", "Saldo", "Origen signo"],
        filas,
        [IZQUIERDA] * 3 + [DERECHA] * 3 + [IZQUIERDA],
    )
    if limite is not None and len(movimientos) > limite:
        texto += f"\n  ... y {len(movimientos) - limite} movimientos más."
    return texto


def _porcentaje(parte: Decimal, total: Decimal) -> str:
    if not total:
        return ""
    return f"{(parte / total * 100):.1f}%".replace(".", ",")


def reporte_por_concepto(movimientos: list[Movimiento]) -> str:
    """En qué se fue la plata y de dónde vino, por concepto."""
    from .clasificacion import por_concepto

    filas_datos = por_concepto(movimientos)
    bloques: list[str] = []

    for tipo, titulo in (("INGRESO", "INGRESOS"), ("EGRESO", "EGRESOS")):
        del_tipo = [f for f in filas_datos if f.tipo == tipo]
        if not del_tipo:
            continue
        total = sum((f.total for f in del_tipo), Decimal("0.00"))
        filas = [
            [
                f.concepto,
                str(f.cantidad),
                formato_cop(f.total),
                _porcentaje(f.total, total),
                formato_cop(f.promedio),
            ]
            for f in del_tipo
        ]
        filas.append(["TOTAL " + titulo, str(sum(f.cantidad for f in del_tipo)),
                      formato_cop(total), "100,0%", ""])
        bloques.append(titulo)
        bloques.append(
            tabla(
                ["Concepto", "Mov.", "Total", "% del total", "Promedio"],
                filas,
                [IZQUIERDA, DERECHA, DERECHA, DERECHA, DERECHA],
            )
        )
        bloques.append("")

    return "\n".join(bloques).rstrip()


def reporte_concepto_y_tercero(movimientos: list[Movimiento]) -> str:
    """El informe que responde "con quién": cada concepto abierto por tercero."""
    from .clasificacion import por_concepto_y_tercero

    filas_datos = por_concepto_y_tercero(movimientos)
    bloques: list[str] = []

    for tipo, titulo in (
        ("INGRESO", "INGRESOS  (de quién entró la plata)"),
        ("EGRESO", "EGRESOS  (a quién se le pagó)"),
    ):
        del_tipo = [f for f in filas_datos if f.tipo == tipo]
        if not del_tipo:
            continue
        total = sum((f.total for f in del_tipo), Decimal("0.00"))

        filas: list[list[str]] = []
        concepto_actual = None
        for fila in del_tipo:
            if fila.concepto != concepto_actual:
                concepto_actual = fila.concepto
                del_concepto = [f for f in del_tipo if f.concepto == concepto_actual]
                subtotal = sum((f.total for f in del_concepto), Decimal("0.00"))
                filas.append(
                    [
                        concepto_actual,
                        str(sum(f.cantidad for f in del_concepto)),
                        formato_cop(subtotal),
                        _porcentaje(subtotal, total),
                        "",
                    ]
                )
            filas.append(
                [
                    f"    {fila.tercero}",
                    str(fila.cantidad),
                    formato_cop(fila.total),
                    "",
                    f"{fila.primera} a {fila.ultima}"
                    if fila.primera != fila.ultima
                    else fila.primera,
                ]
            )

        filas.append(
            [
                "TOTAL",
                str(sum(f.cantidad for f in del_tipo)),
                formato_cop(total),
                "100,0%",
                "",
            ]
        )
        bloques.append(titulo)
        bloques.append(
            tabla(
                ["Concepto / Tercero", "Mov.", "Total", "% del total", "Fechas"],
                filas,
                [IZQUIERDA, DERECHA, DERECHA, DERECHA, IZQUIERDA],
            )
        )
        bloques.append("")

    return "\n".join(bloques).rstrip()


def reporte_terceros(movimientos: list[Movimiento], limite: int | None = None) -> str:
    """Consolidado por tercero, sumando todo lo que se movió con cada uno."""
    from .clasificacion import por_tercero, variantes_por_tercero

    filas_datos = por_tercero(movimientos)
    variantes = variantes_por_tercero(movimientos)
    seleccion = filas_datos if limite is None else filas_datos[:limite]

    filas = []
    for fila in seleccion:
        nombre = fila.tercero
        if nombre in variantes:
            nombre += f"  (+{len(variantes[nombre])} variante(s))"
        filas.append(
            [
                nombre,
                fila.concepto[:38],
                str(fila.cantidad),
                formato_cop(fila.total) if fila.tipo == "INGRESO" else "",
                formato_cop(fila.total) if fila.tipo == "EGRESO" else "",
                ", ".join(fila.bancos)[:24],
            ]
        )

    texto = tabla(
        ["Tercero", "Concepto(s)", "Mov.", "Recibido", "Pagado", "Banco(s)"],
        filas,
        [IZQUIERDA, IZQUIERDA, DERECHA, DERECHA, DERECHA, IZQUIERDA],
    )
    if limite is not None and len(filas_datos) > limite:
        texto += f"\n  ... y {len(filas_datos) - limite} terceros más."
    if variantes:
        texto += (
            "\n\n  Los terceros marcados con variantes agrupan nombres escritos "
            "de distinta forma.\n  Para ver o corregir la agrupación, use el "
            "archivo terceros.csv."
        )
    return texto


def reporte_lectura(consolidado: Consolidado) -> str:
    """Qué se leyó de cada archivo: útil para detectar problemas de lectura."""
    filas = []
    for extracto in consolidado.extractos:
        periodo = "-"
        if extracto.periodo_inicio and extracto.periodo_fin:
            periodo = (
                f"{extracto.periodo_inicio.strftime('%d/%m/%Y')} a "
                f"{extracto.periodo_fin.strftime('%d/%m/%Y')}"
            )
        cuadre = extracto.cuadra
        filas.append(
            [
                os.path.basename(extracto.archivo),
                extracto.banco + ("" if extracto.detectado_automaticamente else " (fijado)"),
                extracto.cuenta or "-",
                periodo,
                str(len(extracto.movimientos)),
                extracto.motor_texto,
                "OK" if cuadre else ("s/saldos" if cuadre is None else "REVISAR"),
            ]
        )
    return tabla(
        ["Archivo", "Banco", "Cuenta", "Periodo", "Mov.", "Lectura", "Cuadre"],
        filas,
        [IZQUIERDA, IZQUIERDA, IZQUIERDA, IZQUIERDA, DERECHA, IZQUIERDA, IZQUIERDA],
    )


def reporte_advertencias(consolidado: Consolidado) -> str:
    lineas: list[str] = []

    # Las advertencias que se repiten en muchos archivos (típicamente las que
    # hablan del motor de lectura) se resumen en un renglón para no tapar las
    # que son propias de un extracto.
    repeticiones: dict[str, int] = {}
    for extracto in consolidado.extractos:
        for advertencia in set(extracto.advertencias):
            repeticiones[advertencia] = repeticiones.get(advertencia, 0) + 1

    umbral = max(3, len(consolidado.extractos))
    generales = {a for a, veces in repeticiones.items() if veces >= umbral}

    for advertencia in sorted(generales):
        lineas.append(f"  En los {repeticiones[advertencia]} archivos:")
        lineas.append(f"    - {advertencia}")

    for extracto in consolidado.extractos:
        propias = [a for a in extracto.advertencias if a not in generales]
        if not propias:
            continue
        lineas.append(f"  {os.path.basename(extracto.archivo)}:")
        for advertencia in propias:
            lineas.append(f"    - {advertencia}")
    for ruta, error in consolidado.errores:
        lineas.append(f"  {os.path.basename(ruta)}: NO SE PUDO LEER -> {error}")
    if consolidado.duplicados:
        lineas.append(
            f"  Se descartaron {len(consolidado.duplicados)} movimientos "
            "repetidos entre archivos."
        )
    return "\n".join(lineas)


# ---------------------------------------------------------------------------
# Archivos de salida
# ---------------------------------------------------------------------------

def etiqueta_archivo(texto: str) -> str:
    """Convierte un nombre en algo válido para un archivo de Windows.

    "Banco de Bogotá" -> "Banco_de_Bogota"
    """
    from .normalizacion import sin_acentos

    limpio = sin_acentos(texto or "").strip()
    limpio = re.sub(r"[^\w\s-]", "", limpio)
    limpio = re.sub(r"\s+", "_", limpio)
    return limpio or "sin_nombre"


def _valor_csv(valor: object) -> object:
    """Los decimales se escriben con coma para que Excel en español los lea."""
    if isinstance(valor, Decimal):
        return formato_cop(valor)
    return valor


def escribir_csv(ruta: str, campos: list[str], filas: list[dict[str, object]]) -> str:
    os.makedirs(os.path.dirname(os.path.abspath(ruta)) or ".", exist_ok=True)
    # utf-8-sig para que Excel reconozca las tildes al abrir el archivo.
    with open(ruta, "w", newline="", encoding="utf-8-sig") as archivo:
        escritor = csv.DictWriter(archivo, fieldnames=campos, delimiter=";")
        escritor.writeheader()
        for fila in filas:
            escritor.writerow({c: _valor_csv(fila.get(c, "")) for c in campos})
    return ruta


def filas_resumen(resumenes: list[ResumenMensual]) -> list[dict[str, object]]:
    filas = []
    for resumen in sorted(resumenes, key=lambda r: (r.periodo, r.banco)):
        filas.append(
            {
                "periodo": resumen.periodo,
                "mes": periodo_legible(resumen.periodo),
                "banco": resumen.banco,
                "cuenta": resumen.cuenta or "",
                "movimientos": resumen.cantidad_movimientos,
                "saldo_inicial": resumen.saldo_inicial,
                "ingresos": resumen.ingresos,
                "egresos": resumen.egresos,
                "neto": resumen.neto,
                "saldo_final": resumen.saldo_final,
                "saldo_final_extracto": resumen.saldo_final_extracto,
                "diferencia_cuadre": resumen.diferencia,
                "cuadre": _marca_cuadre(resumen),
                "sin_clasificar": resumen.movimientos_sin_clasificar,
                "clasificados_por_texto": resumen.movimientos_por_palabras,
            }
        )
    return filas


CAMPOS_RESUMEN = [
    "periodo", "mes", "banco", "cuenta", "movimientos", "saldo_inicial",
    "ingresos", "egresos", "neto", "saldo_final", "saldo_final_extracto",
    "diferencia_cuadre", "cuadre", "sin_clasificar", "clasificados_por_texto",
]

CAMPOS_MOVIMIENTOS = [
    "banco", "cuenta", "fecha", "periodo", "descripcion", "concepto", "tercero",
    "referencia", "documento", "tipo", "ingreso", "egreso", "valor", "saldo",
    "confianza", "archivo",
]

CAMPOS_CONCEPTOS = ["tipo", "concepto", "movimientos", "total", "porcentaje",
                    "promedio"]

CAMPOS_CONCEPTO_TERCERO = ["tipo", "concepto", "tercero", "movimientos", "total",
                           "bancos", "primera_fecha", "ultima_fecha"]

CAMPOS_TERCEROS = ["tercero", "conceptos", "movimientos", "recibido", "pagado",
                   "neto", "bancos", "primera_fecha", "ultima_fecha",
                   "variantes_agrupadas"]


def filas_conceptos(movimientos: list[Movimiento]) -> list[dict[str, object]]:
    from .clasificacion import por_concepto

    filas_datos = por_concepto(movimientos)
    totales = {
        tipo: sum((f.total for f in filas_datos if f.tipo == tipo), Decimal("0.00"))
        for tipo in ("INGRESO", "EGRESO")
    }
    return [
        {
            "tipo": f.tipo,
            "concepto": f.concepto,
            "movimientos": f.cantidad,
            "total": f.total,
            "porcentaje": _porcentaje(f.total, totales[f.tipo]),
            "promedio": f.promedio,
        }
        for f in filas_datos
    ]


def filas_concepto_tercero(movimientos: list[Movimiento]) -> list[dict[str, object]]:
    from .clasificacion import por_concepto_y_tercero

    return [
        {
            "tipo": f.tipo,
            "concepto": f.concepto,
            "tercero": f.tercero,
            "movimientos": f.cantidad,
            "total": f.total,
            "bancos": ", ".join(f.bancos),
            "primera_fecha": f.primera,
            "ultima_fecha": f.ultima,
        }
        for f in por_concepto_y_tercero(movimientos)
    ]


def filas_terceros(movimientos: list[Movimiento]) -> list[dict[str, object]]:
    from .clasificacion import por_tercero, variantes_por_tercero

    variantes = variantes_por_tercero(movimientos)
    filas = []
    for fila in por_tercero(movimientos):
        recibido = sum(
            (m.ingreso for m in movimientos if (m.tercero or "") == fila.tercero),
            Decimal("0.00"),
        )
        pagado = sum(
            (m.egreso for m in movimientos if (m.tercero or "") == fila.tercero),
            Decimal("0.00"),
        )
        filas.append(
            {
                "tercero": fila.tercero,
                "conceptos": fila.concepto,
                "movimientos": fila.cantidad,
                "recibido": recibido,
                "pagado": pagado,
                "neto": (recibido - pagado),
                "bancos": ", ".join(fila.bancos),
                "primera_fecha": fila.primera,
                "ultima_fecha": fila.ultima,
                "variantes_agrupadas": " | ".join(
                    sorted(variantes.get(fila.tercero, []))
                ),
            }
        )
    return filas


def exportar(
    carpeta: str,
    *,
    consolidado: Consolidado,
    resumenes: list[ResumenMensual],
    totales: list[TotalBanco],
    formatos: list[str],
    prefijo: str | None = None,
) -> list[str]:
    """Escribe los reportes en disco. Devuelve las rutas generadas."""
    os.makedirs(carpeta, exist_ok=True)
    marca = prefijo or datetime.now().strftime("%Y%m%d_%H%M")
    generados: list[str] = []

    clasificados = any(m.concepto for m in consolidado.movimientos)

    if "csv" in formatos:
        generados.append(
            escribir_csv(
                os.path.join(carpeta, f"{marca}_resumen_mensual.csv"),
                CAMPOS_RESUMEN,
                filas_resumen(resumenes),
            )
        )
        generados.append(
            escribir_csv(
                os.path.join(carpeta, f"{marca}_movimientos.csv"),
                CAMPOS_MOVIMIENTOS,
                [m.como_fila() for m in consolidado.movimientos],
            )
        )
        if clasificados:
            generados.append(
                escribir_csv(
                    os.path.join(carpeta, f"{marca}_por_concepto.csv"),
                    CAMPOS_CONCEPTO_TERCERO,
                    filas_concepto_tercero(consolidado.movimientos),
                )
            )
            generados.append(
                escribir_csv(
                    os.path.join(carpeta, f"{marca}_por_tercero.csv"),
                    CAMPOS_TERCEROS,
                    filas_terceros(consolidado.movimientos),
                )
            )

    if "xlsx" in formatos:
        ruta = os.path.join(carpeta, f"{marca}_conciliacion.xlsx")
        generado = escribir_excel(ruta, consolidado, resumenes, totales)
        if generado:
            generados.append(generado)

    return generados


def escribir_excel(
    ruta: str,
    consolidado: Consolidado,
    resumenes: list[ResumenMensual],
    totales: list[TotalBanco],
) -> str | None:
    """Genera un Excel con una hoja por reporte. Requiere openpyxl."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        return None

    libro = Workbook()
    formato_numero = '#,##0.00;[Red]-#,##0.00'

    def agregar_hoja(titulo: str, campos: list[str], filas: list[dict[str, object]]):
        hoja = libro.create_sheet(titulo)
        hoja.append([c.replace("_", " ").title() for c in campos])
        for celda in hoja[1]:
            celda.font = Font(bold=True, color="FFFFFF")
            celda.fill = PatternFill("solid", start_color="2F5597")
            celda.alignment = Alignment(horizontal="center")
        for fila in filas:
            hoja.append([fila.get(campo, "") for campo in campos])
        for indice, campo in enumerate(campos, start=1):
            letra = get_column_letter(indice)
            ancho = max(
                [len(campo) + 2]
                + [len(str(f.get(campo, ""))) + 2 for f in filas[:400]]
            )
            hoja.column_dimensions[letra].width = min(ancho, 48)
            if any(
                clave in campo
                for clave in ("saldo", "ingreso", "egreso", "valor", "neto",
                              "diferencia")
            ):
                for celda in hoja[letra][1:]:
                    celda.number_format = formato_numero
        hoja.freeze_panes = "A2"
        return hoja

    libro.remove(libro.active)

    agregar_hoja("Resumen mensual", CAMPOS_RESUMEN, filas_resumen(resumenes))

    if any(m.concepto for m in consolidado.movimientos):
        agregar_hoja(
            "Por concepto", CAMPOS_CONCEPTOS, filas_conceptos(consolidado.movimientos)
        )
        agregar_hoja(
            "Concepto y tercero",
            CAMPOS_CONCEPTO_TERCERO,
            filas_concepto_tercero(consolidado.movimientos),
        )
        agregar_hoja(
            "Terceros", CAMPOS_TERCEROS, filas_terceros(consolidado.movimientos)
        )
    agregar_hoja(
        "Totales por banco",
        ["banco", "cuentas", "meses", "movimientos", "saldo_inicial", "ingresos",
         "egresos", "neto", "saldo_final"],
        [
            {
                "banco": t.banco,
                "cuentas": ", ".join(t.cuentas),
                "meses": len(t.meses),
                "movimientos": t.cantidad_movimientos,
                "saldo_inicial": t.saldo_inicial,
                "ingresos": t.ingresos,
                "egresos": t.egresos,
                "neto": t.neto,
                "saldo_final": t.saldo_final,
            }
            for t in totales
        ],
    )
    agregar_hoja(
        "Movimientos",
        CAMPOS_MOVIMIENTOS,
        [m.como_fila() for m in consolidado.movimientos],
    )

    filas_archivos = []
    for extracto in consolidado.extractos:
        filas_archivos.append(
            {
                "archivo": os.path.basename(extracto.archivo),
                "banco": extracto.banco,
                "cuenta": extracto.cuenta or "",
                "desde": extracto.periodo_inicio,
                "hasta": extracto.periodo_fin,
                "movimientos": len(extracto.movimientos),
                "saldo_inicial": extracto.saldo_inicial,
                "saldo_final": extracto.saldo_final,
                "diferencia": extracto.diferencia_cuadre,
                "lectura": extracto.motor_texto,
                "advertencias": " | ".join(extracto.advertencias),
            }
        )
    agregar_hoja(
        "Archivos leidos",
        ["archivo", "banco", "cuenta", "desde", "hasta", "movimientos",
         "saldo_inicial", "saldo_final", "diferencia", "lectura", "advertencias"],
        filas_archivos,
    )

    os.makedirs(os.path.dirname(os.path.abspath(ruta)) or ".", exist_ok=True)
    libro.save(ruta)
    return ruta
